from tkinter import *
from tkinter import ttk,messagebox
import random
import openpyxl
import xlrd
from PIL import ImageTk,Image
from tkinter import filedialog
import pandas as pd
from openpyxl import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from playsound import playsound
import pygame
import threading
import time
import xlsxwriter
from xlrd import *

import numpy as np
import matplotlib.pyplot as plt
from fpdf import FPDF


global Green
global Red
global Yellow
global Blue
global watch
global count
global tmp
global gameover
global timer_count
global score
global tmp_stt
global stt

score = 0
timer_count = 0
watch = True
gameover = False
stt = []
tmp = len(stt)
tmp_stt = []
global user_r_temp
global user_p_tmp
user_r_temp=0
user_p_tmp=0
global helpvair
helpvair=0
global helpvairp
helpvairp=0
global num1
num1=0
global num2
num2=0
global count1_reg
count1_reg=0
global count2_prob
count2_prob=0
global num1_admin
num1_admin = 0
global num2_admin
num2_admin = 0

class Register:
    pygame.mixer.init()

    def __init__(self,root):
        self.root=root
        self.root.title('Simon Game')
        self.root.geometry('2048x2048')
        self.login()
        # adding icon
        #root.iconbitmap('simon_icon.a84d8b33-ba56-4537-8859-3c3b04b95039')
    def register(self):
        #self.root.config(bg="black") --------why this not working???
        #bg image
        self.bg=ImageTk.PhotoImage(file="img3.jpg")
        bg=Label(self.root,image=self.bg).place(x=250,y=0,relwidth=1,relheight=1)
        # left image
        self.left = ImageTk.PhotoImage(file="as.png")
        left= Label(self.root, image=self.left,bg="gray").place(x=80, y=100, width=472, height=512)
        #register fram
        frame1=Frame(self.root,bg="white")
        frame1.place(x=550,y=100,width=700,height=512)
        title=Label(frame1,text="REGISTRATION FORM",font=("times new roman",20,"bold"),bg="white",fg="red").place(x=50,y=30)
        username = Label(frame1, text="UserName", font=("times new roman", 15, "bold"), bg="white",fg="black").place(x=50, y=100)
        self.txt_username=Entry(frame1,font=("times new roman", 15),bg="lightgray")
        self.txt_username.place(x=50,y=130,width=250)
        age = Label(frame1, text="Age", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=50, y=170)
        self.txt_age = Entry(frame1, font=("times new roman", 15), bg="lightgray")
        self.txt_age.place(x=50, y=200, width=250)
        usertype = Label(frame1, text="User Type", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=50, y=240)
        # adding radio button to the form
        self.r = IntVar()

        Radiobutton(frame1, text="Regular user", variable=self.r, value=1,font=("times new roman", 15), bg="white").place(x=50, y=300)
        Radiobutton(frame1, text="User with vision problem", variable=self.r, value=2,font=("times new roman", 15), bg="white").place(x=200, y=300)
        '''Radiobutton(frame1, text="Admin", variable=self.r, value=3,font=("times new roman", 15), bg="white").place(x=450, y=300)'''
        #button for register
        btn=Button(frame1,text="Register me -->",font=("times new roman", 15, "bold"),bg="lightblue",width=30,cursor="hand2",command=self.register_data).place(x=50 ,y=400)
        #button for log in
        btn_login = Button(self.root, text="Log In ", font=("times new roman", 20), bg="lightblue", width=10,cursor="hand2",command=self.login).place(x=250, y=560)

    def register_data(self):
        # if you clicked register then:
        if self.txt_username.get()=="" or self.txt_age.get()=="" or (self.r.get()!=1 and self.r.get()!=2):
            messagebox.showerror("ERROR","All Fields Are Required",parent=self.root)
        else:
            wb=openpyxl.load_workbook('good.xlsx')
            ws=wb['Sheet1']
            users_name=ws.iter_rows(min_row=2)
            flag=0
            for cell in users_name:
                if(cell[0].value==self.txt_username.get() or self.txt_username.get()=='admin'):
                    messagebox.showerror("ERROR", "User Already Exist", parent=self.root)
                    self.txt_username.delete(0, END)
                    self.txt_age.delete(0, END)
                    flag=1

            if(flag==0):
                # if the user enter all the data then we will save it at excel page
                if (self.r.get() == 1):
                    ustype = "Regular user"
                if (self.r.get() == 2):
                    ustype = "User with vision problem"
                # ---------
                path = 'good.xlsx'
                df1 = pd.read_excel(path)
                SeriesA = df1['user']  # what in the quote is the name of the line at the excel file
                SeriesB = df1['Age']
                SeriesC = df1['UserType']
                A = pd.Series(self.txt_username.get())
                B = pd.Series(self.txt_age.get())
                C = pd.Series(ustype)
                SeriesA = SeriesA.append(A)
                SeriesB = SeriesB.append(B)
                SeriesC = SeriesC.append(C)
                df2 = pd.DataFrame({"user": SeriesA, "Age": SeriesB, "UserType": SeriesC})
                df2.to_excel(path, index=False)
                #****************************************
                #****************************************
                #save the name at the rating excel page
                wb_rate_name= openpyxl.load_workbook('rating.xlsx')
                sheet_rate_name = wb_rate_name['Sheet1']
                for i, _ in enumerate(iter(bool, True), start=1):
                    if (sheet_rate_name.cell(row=i, column=1).value is None):
                        sheet_rate_name.cell(row=i, column=1).value = self.txt_username.get()
                        wb_rate_name.save('rating.xlsx')
                        break
                wb_rate_name.close()
                # ****************************************
                # ****************************************
                if (self.r.get() == 1):  #user1 - save the name of user from type 1 also at the user1 excel sheet
                    wb1 = openpyxl.load_workbook('user1.xlsx')
                    sheet = wb1['Sheet1']
                    for i, _ in enumerate(iter(bool,True),start=1):
                        if(sheet.cell(row=1,column=i).value is None):
                            #print(sheet.cell(row=1,column=i).value)
                            sheet.cell(row=1, column=i).value = self.txt_username.get()
                            wb1.save('user1.xlsx')
                            break
                    wb1.close()
                        #else:
                         #   print("somthing hereeee")

                if (self.r.get() == 2):  # user2 - save the name of user from type 2 also at the user2 excel sheet
                    wb2 = openpyxl.load_workbook('user2.xlsx')
                    sheet2 = wb2['Sheet1']
                    for i,_ in enumerate(iter(bool,True),start=1):
                        if(sheet2.cell(row=1,column=i).value is None):
                            sheet2.cell(row=1, column=i).value = self.txt_username.get()
                            wb2.save('user2.xlsx')
                            break
                    wb2.close()




                #$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
                self.txt_username.delete(0, END)
                self.txt_age.delete(0, END)
                messagebox.showinfo("WELCOME ", "Register Successful", parent=self.root)
            #wb.save('good.xlsx')
            #wb.close()
    def login(self):
        self.bg1 = ImageTk.PhotoImage(file="img3.jpg")
        bg1= Label(self.root, image=self.bg1).place(x=0, y=0, relwidth=1, relheight=1)
        frame2 = Frame(self.root, bg="white")
        frame2.place(x=200, y=100, width=500, height=412)
        title1 = Label(frame2, text="Login Here", font=("times new roman", 32, "bold"), bg="white",fg="red").place(x=50, y=30)
        username_log = Label(frame2, text="UserName", font=("times new roman", 20, "bold"), bg="white", fg="steelblue").place(x=50, y=130)
        self.txt_username_log = Entry(frame2, font=("times new roman", 15), bg="lightgray")
        self.txt_username_log.place(x=50, y=180, width=200)
        # button for log in
        btn1 = Button(frame2, text="Login", font=("times new roman", 15, "bold"), bg="steelblue",fg="white", width=15,cursor="hand2",command=self.log_button).place(x=50, y=280)
        btn2 = Button(frame2, text="Not Registered? Regist Now", font=("times new roman", 10, "bold"), bg="white", fg="black", bd=0,width=25,cursor="hand2",command=self.register).place(x=50, y=240)
    def log_button(self):
        if (self.txt_username_log.get()==""):
            messagebox.showerror("Error", "You Must Enter The UserName", parent=self.root)
        else:
            wb = openpyxl.load_workbook('good.xlsx')
            ws = wb['Sheet1']
            users_name = ws.iter_rows(min_row=2)
            flag = 0
            for cell in users_name:
                if (cell[0].value == self.txt_username_log.get()):
                    flag = 1
            if(flag==0 and self.txt_username_log.get() != 'admin'):
                messagebox.showerror("Error", "You Entered an incorrect UserName ", parent=self.root)
                self.txt_username_log.delete(0, END) # i have to save the data from the login page in help vairble to save the data with it
            if(flag==1):
                self.The_User = self.txt_username_log.get() #@@@@@@@@@@
                self.txt_username_log.delete(0, END)
                self.menu()
            if (self.txt_username_log.get() == 'admin'):
                self.admin()


root=Tk()
obj=Register(root)
root.mainloop()