
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from tkinter import *
from tkinter import messagebox
root = Tk()
#open the file we save the data in it
wb=load_workbook('users.xlsx')
ws=wb.active
#save the data in the file and check if the name is used
def Register():
    username=Username.get()
    age=Age.get()
    password=Password.get()
    if username:
        used = False
        for cell in ws['A']:
            if username == cell.value:
                messagebox.showinfo("", "The name is used")
                used=True
        if username==''or age == ''or password=='':
            messagebox.showinfo("", "Enter something in the blank ")
            username = Username.get()
            age = Age.get()
            used = True

        if not used:
            ws.cell(column=1, row=ws.max_row + 1, value=username)
            ws.cell(column=2, row=ws.max_row , value=password)
            ws.cell(column=3, row=ws.max_row , value=age)
            messagebox.showinfo("", "Register success")

        wb.save("users.xlsx")

#the function  that open the Register_window
def Register_window():
    root.title("Register")
    root.geometry("300x300")
    global Username
    global Age
    global Password
    #username
    Label(root, text="Username:").place(x=20, y=20)
    Username = Entry(root, bd=5)
    Username.place(x=140, y=20)
    # password
    Label(root, text="Password:").place(x=20, y=60)
    Password = Entry(root, bd=5)
    Password.place(x=140, y=60)
    #age
    Label(root, text="Age:").place(x=20, y=100)
    Age = Entry(root, bd=5)
    Age.place(x=140, y=100)

    Register_window()
Button(root,text="Register",command=Register,heigh=2,width=12,bd=3).place(x=30,y=200)
root.mainloop()
